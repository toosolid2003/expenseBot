from datetime import datetime
import openpyxl
from os import path
import zipfile
import time
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import (Mail, Attachment, FileContent, FileName, FileType, Disposition, ContentId)
import base64
import sqlite3
from logger.logger import logger

class ExpenseReport:
    
    def __init__(self, user):
        self.date = datetime.today()
        self.timestamp = int(time.time())
        self.format = '.xlsx'
        self.template = '/var/www/expenseBot/exports/exportTemplate.xlsx'
        self.expenseList = None
        self.user = user
        self.receiptsPath = None
        self.reportPath = None

    def getExpenses(self, db, date_exp=None):
        """
        Gets the expenses from the database. Returns a series of Row objects.
        Inputs: database name/path, [optional] date of first expense (YYYY-MM-DD)
        Output: list of sqlite3.Row objects
        """

        #Setting up the connection with the db
        logger.debug('[*] Connecting to the db')
        conn = sqlite3.connect(db, check_same_thread=False)
        conn.row_factory = sqlite3.Row

        #Building the query
        query = 'SELECT * FROM items WHERE user=?'
        data = (self.user,)

        if date_exp != None:
            query += ' AND date_expense>=?'
            data = (self.user, date_exp)

        #Launching the query
        logger.debug('[*] Executing the query')
        c = conn.cursor()
        c.execute(query, data)
        self.expenseList = c.fetchall()
        conn.close()

        return True

    def generateXls(self):
        
        """Generates an expense report. 

        Input: list of expenses
        Output: absolute path to the expense report"""

        #Loading the excel template
        logger.debug('[*] Loading the workbook template')
        wb = openpyxl.load_workbook(filename=self.template)
        ws = wb.active
        row = 19
        col = 2


        #Insert all the expenses in the Excel template
        logger.debug('[*] Filling the spreadsheet')
        for expense in self.expenseList:
            ws.cell(row, col).value = expense['reason']
            ws.cell(row, col + 1).value = expense['amount']
            ws.cell(row, col + 2).value = expense['currency']
            ws.cell(row, col + 3).value = expense['date_expense']
            ws.cell(row, col + 4).value = expense['typex']
            ws.cell(row, col + 5).value = path.basename(expense['receipt'])
            row += 1
        
        #Adding the total amount
        row += 2
        ws.cell(row, col ).value = 'TOTAL'

        formula = "=SUM(C19:C" + str(row-1) + ")" 
        ws.cell(row, col + 1).value = formula


        #Saving the file
        logger.debug('[*] Saving the file on disk')
        self.reportPath = '/var/www/expenseBot/exports/' + self.user + '/' + 'expenseReport_' + str(self.timestamp) + self.format
        wb.save(self.reportPath)

        return self.reportPath
    
    def receiptZip(self):
        '''Gets the receipts for the expense report of the object.'''

        #Flename creation
        logger.debug('[*] Creating filename for zip file')
        self.receiptsPath = '/var/www/expenseBot/exports/' + self.user + '/' + 'receiptsExports_' + str(self.timestamp) + '.zip'

        #Creating the zipfile
        logger.debug('[*] Filling the zip archive with receipts')
        with zipfile.ZipFile(self.receiptsPath,'x') as myzip:
            for expense in self.expenseList:
                myzip.write(expense['receipt'], path.basename(expense['receipt']))   #The path of the receipt is in 5th position in the expense list
    
        #Return the file path
        return self.receiptsPath
 
    
    def sendMail(self, email):
        '''Sends the expense report with corresponding receipts using the SendGrid platform
        
        Input: destination email, path pf the expense report, path of the archive containing the receipts
        Output: feedback code sent by SendGrid'''

        """
        Send an email with attatchment to specified recipient with sendGrid.
        Input: email of sender, email of recipient, expense report file path and receipts file path
        Output: reponse code from SendGrid

        Make sure that the email of sender has been verified on sendGrid.
        """


        #Initiate constants and objects

        API = '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'
        sg = SendGridAPIClient(api_key=API)

        #Creating Mail object
        logger.debug('[*] Initiating mail object')
        message = Mail(
            from_email = 'support@expensebot.net',
            to_emails= email,
            subject = 'Your expense report',
            html_content = '<strong>Here is your expense report, with receipts, from expenseBot.net</strong>'
            )

        #Creating expense report attachment
        logger.debug('[*] Creating the expense report attachment')
        with open(self.reportPath, 'rb') as f:
            data = f.read()
            f.close()

        encoded_file = base64.b64encode(data).decode()

        expenseReportAttachment = Attachment(
            FileContent(encoded_file),
            FileName('yourExport.xlsx'),
            FileType('text/csv'),
            Disposition('attachment')
            )

        message.add_attachment(expenseReportAttachment)

        #Creating receipts attachment
        logger.debug('[*] Creating the receipts attachment')
        with open(self.receiptsPath, 'rb') as file:
            data = file.read()
            file.close()
        encoded_file = base64.b64encode(data).decode()

        receiptsAttachment = Attachment(
            FileContent(encoded_file),
            FileName('receipts.zip'),
            FileType('text/zip'),
            Disposition('attachment')
        )
        message.add_attachment(receiptsAttachment)

        #Sending the email
        logger.debug('[*] Sending the email')
        response = sg.send(message)

        return response.status_code