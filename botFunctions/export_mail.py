# coding: utf-8
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import (Mail, Attachment, FileContent, FileName, FileType, Disposition, ContentId)
import base64
import openpyxl
from botClasses.classes import DBHelper
from os import path

#variables: 
# - filepath
# - email of sender
# - email of recipient

def sendExport(emailSender, emailRecipient, expenseReport, receipts):
    """
    Send an email with attatchment to specified recipient with sendGrid.
    Input: email of sender, email of recipient, expense report file path and receipts file path
    Output: reponse code

    Make sure that the email of sender has been verified on sendGrid.
    """
    #Initiate constants and objects

    API = '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'
    sg = SendGridAPIClient(api_key=API)

    #Creating Mail object
    message = Mail(
        from_email = emailSender,
        to_emails= emailRecipient,
        subject = 'Your csv export',
        html_content = '<strong>Here is your csv export from expenseBot.net</strong>'
        )

    #Creating expense report attachment
    with open(expenseReport, 'rb') as f:
        data = f.read()
        f.close()

    encoded_file = base64.b64encode(data).decode()

    expenseReportAttachment = Attachment(
        FileContent(encoded_file),
        FileName('yourExport.csv'),
        FileType('text/csv'),
        Disposition('attachment')
        )

    message.add_attachment(expenseReportAttachment)

    #Creating receipts attachment
    with open(receipts, 'rb') as file:
        data = file.read()
        file.close()
    encoded_file = base64.b64encode(data).decode()

    receiptsAttachment = Attachment(
        FileContent(encoded_file),
        FileName('receipts.zip'),
        FileType('text/zip'),
        Disposition('attachment')
    )
    message.add_attachment(receiptsAttachment)

    #Sending the email

    response = sg.send(message)

    return response

def formatExport(expenses):
    '''Format the list of expenses in an Excel spreadsheet.
    Input: list of expenses to be formatted
    Output: path of the saved Excel spreadsheet'''

    #Loading the excel template
    wb = openpyxl.load_workbook(filename='/var/www/expenseBot/exports/exportTemplate.xlsx')
    ws = wb.active
    row = 19
    col = 2

    #Removing the absolute path from the receipt column
    #Converting the result into a list of lists to allow reformatting later on
    exp = []
    for elt in expenses:
        exp.append(list(elt))

    #Reformating the 'receipts' string to remove the abosulte path
    for elt in exp:
        elt[5] = path.basename(elt[5])


    #Format the receipt column to remove the whole path
    for reason, amount, currency, date, typex, receipt in exp:
        ws.cell(row, col).value = reason
        ws.cell(row, col + 1).value = amount
        ws.cell(row, col + 2).value = currency
        ws.cell(row, col + 3).value = date
        ws.cell(row, col + 4).value = typex
        ws.cell(row, col + 5).value = receipt
        row += 1

    #Saving the file
    file = '/var/www/expenseBot/exports/yourExport.xlsx'
    wb.save(file)